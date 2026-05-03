#!/usr/bin/env python3
"""
QR Product Generator - Enhanced Font/Logo/Border Edition

Features
- GUI mode for single and batch QR generation
- PNG export
- Optional center logo
- Optional outer border around final image
- Optional caption under QR using the entered filename
- Custom font file selection for Persian/Arabic or any other font
- Caption font size and color controls
- Optional border on/off
- Clipboard paste helpers for Windows/Tkinter
- CLI mode for single or batch generation

Requires:
    pip install qrcode pillow
Optional for better Persian caption rendering:
    pip install arabic-reshaper python-bidi
"""

from __future__ import annotations

import argparse
import csv
import json
import re
import sys
import zipfile
import html
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable, List, Optional
from urllib.parse import urlparse
from urllib.request import Request, urlopen

import qrcode
from qrcode.constants import ERROR_CORRECT_L, ERROR_CORRECT_M, ERROR_CORRECT_Q, ERROR_CORRECT_H
from PIL import Image, ImageDraw, ImageFont, ImageTk

try:
    import tkinter as tk
    from tkinter import colorchooser, filedialog, messagebox, ttk
    from tkinter.scrolledtext import ScrolledText
except Exception:  # pragma: no cover
    tk = None
    colorchooser = None
    filedialog = None
    messagebox = None
    ttk = None
    ScrolledText = None

try:  # Optional, for nicer Persian/Arabic text rendering
    import arabic_reshaper
    from bidi.algorithm import get_display
except Exception:  # pragma: no cover
    arabic_reshaper = None
    get_display = None


ERROR_LEVELS = {
    "L": ERROR_CORRECT_L,
    "M": ERROR_CORRECT_M,
    "Q": ERROR_CORRECT_Q,
    "H": ERROR_CORRECT_H,
}


@dataclass
class QRJob:
    name: str
    url: str
    code: str = ""


@dataclass
class RenderOptions:
    box_size: int = 10
    border: int = 4
    error_level: str = "H"
    fill_color: str = "black"
    back_color: str = "white"
    logo_path: Optional[str] = None
    logo_scale_percent: int = 18
    show_frame: bool = True
    frame_width: int = 12
    show_caption: bool = True
    caption_font_path: Optional[str] = None
    caption_font_size: int = 28
    caption_color: str = "#000000"
    caption_direction: str = "auto"


def is_valid_url(value: str) -> bool:
    value = value.strip()
    if not value:
        return False
    parsed = urlparse(value)
    return parsed.scheme in {"http", "https"} and bool(parsed.netloc)


def sanitize_filename(name: str, fallback: str = "qr_code") -> str:
    name = (name or "").strip()
    if not name:
        name = fallback
    name = re.sub(r"[\\/:*?\"<>|]+", "_", name)
    name = re.sub(r"\s+", "_", name)
    name = re.sub(r"_+", "_", name).strip("._")
    return name[:120] or fallback


def contains_rtl_chars(text: str) -> bool:
    return any(
        "؀" <= ch <= "ۿ" or
        "ݐ" <= ch <= "ݿ" or
        "ࢠ" <= ch <= "ࣿ"
        for ch in (text or "")
    )


def normalize_direction_value(direction: str) -> str:
    direction = (direction or "").strip().lower()
    mapping = {
        "auto": "auto",
        "rtl": "rtl",
        "ltr": "ltr",
        "خودکار": "auto",
        "راست‌به‌چپ": "rtl",
        "راست به چپ": "rtl",
        "چپ‌به‌راست": "ltr",
        "چپ به راست": "ltr",
    }
    return mapping.get(direction, "auto")


def resolve_caption_direction(text: str, direction: str = "auto") -> str:
    direction = normalize_direction_value(direction)
    if direction == "auto":
        return "rtl" if contains_rtl_chars(text) else "ltr"
    return direction


def normalize_persian_characters(text: str) -> str:
    """
    Normalize Arabic codepoints commonly seen in Persian text to Persian forms.
    This prevents mixed Arabic/Persian glyphs such as Arabic Yeh (ي) in captions.
    """
    if not text:
        return ""
    translation_table = str.maketrans({
        "ي": "ی",  # Arabic Yeh -> Persian Yeh
        "ى": "ی",  # Alef Maqsura -> Persian Yeh
        "ك": "ک",  # Arabic Kaf -> Persian Keheh
    })
    return text.translate(translation_table)


def prepare_caption_text(text: str, direction: str = "auto") -> str:
    text = normalize_persian_characters((text or "").strip())
    if not text:
        return ""

    resolved_direction = resolve_caption_direction(text, direction)
    if resolved_direction == "ltr":
        return text

    if arabic_reshaper and get_display:
        try:
            return get_display(arabic_reshaper.reshape(text))
        except Exception:
            pass

    # Fallback when bidi/shaping libraries are not installed.
    # 1) Convert Arabic/Persian letters to presentation forms so glyphs can join.
    # 2) Reverse for visual RTL order in PIL's basic renderer.
    return "‏" + shape_arabic_presentation_forms(text)[::-1]


def shape_arabic_presentation_forms(text: str) -> str:
    """
    Lightweight Arabic/Persian shaping fallback.
    Returns text with Arabic Presentation Forms so connected glyphs appear
    even without external reshaper libraries.
    """
    forms: dict[str, tuple[str, Optional[str], Optional[str], Optional[str]]] = {
        "ا": ("ﺍ", "ﺎ", None, None),
        "آ": ("ﺁ", "ﺂ", None, None),
        "أ": ("ﺃ", "ﺄ", None, None),
        "إ": ("ﺇ", "ﺈ", None, None),
        "ب": ("ﺏ", "ﺐ", "ﺑ", "ﺒ"),
        "پ": ("ﭖ", "ﭗ", "ﭘ", "ﭙ"),
        "ت": ("ﺕ", "ﺖ", "ﺗ", "ﺘ"),
        "ث": ("ﺙ", "ﺚ", "ﺛ", "ﺜ"),
        "ج": ("ﺝ", "ﺞ", "ﺟ", "ﺠ"),
        "چ": ("ﭺ", "ﭻ", "ﭼ", "ﭽ"),
        "ح": ("ﺡ", "ﺢ", "ﺣ", "ﺤ"),
        "خ": ("ﺥ", "ﺦ", "ﺧ", "ﺨ"),
        "د": ("ﺩ", "ﺪ", None, None),
        "ذ": ("ﺫ", "ﺬ", None, None),
        "ر": ("ﺭ", "ﺮ", None, None),
        "ز": ("ﺯ", "ﺰ", None, None),
        "ژ": ("ﮊ", "ﮋ", None, None),
        "س": ("ﺱ", "ﺲ", "ﺳ", "ﺴ"),
        "ش": ("ﺵ", "ﺶ", "ﺷ", "ﺸ"),
        "ص": ("ﺹ", "ﺺ", "ﺻ", "ﺼ"),
        "ض": ("ﺽ", "ﺾ", "ﺿ", "ﻀ"),
        "ط": ("ﻁ", "ﻂ", "ﻃ", "ﻄ"),
        "ظ": ("ﻅ", "ﻆ", "ﻇ", "ﻈ"),
        "ع": ("ﻉ", "ﻊ", "ﻋ", "ﻌ"),
        "غ": ("ﻍ", "ﻎ", "ﻏ", "ﻐ"),
        "ف": ("ﻑ", "ﻒ", "ﻓ", "ﻔ"),
        "ق": ("ﻕ", "ﻖ", "ﻗ", "ﻘ"),
        "ک": ("ﮎ", "ﮏ", "ﮐ", "ﮑ"),
        "ك": ("ﻙ", "ﻚ", "ﻛ", "ﻜ"),
        "گ": ("ﮒ", "ﮓ", "ﮔ", "ﮕ"),
        "ل": ("ﻝ", "ﻞ", "ﻟ", "ﻠ"),
        "م": ("ﻡ", "ﻢ", "ﻣ", "ﻤ"),
        "ن": ("ﻥ", "ﻦ", "ﻧ", "ﻨ"),
        "و": ("ﻭ", "ﻮ", None, None),
        "ؤ": ("ﺅ", "ﺆ", None, None),
        "ه": ("ﻩ", "ﻪ", "ﻫ", "ﻬ"),
        "ة": ("ﺓ", "ﺔ", None, None),
        "ی": ("ﯼ", "ﯽ", "ﯾ", "ﯿ"),
        "ي": ("ﻱ", "ﻲ", "ﻳ", "ﻴ"),
        "ئ": ("ﺉ", "ﺊ", "ﺋ", "ﺌ"),
    }

    chars = list(text)
    shaped: list[str] = []

    def can_join_prev(ch: str) -> bool:
        f = forms.get(ch)
        return bool(f and (f[1] or f[3]))

    def can_join_next(ch: str) -> bool:
        f = forms.get(ch)
        return bool(f and (f[2] or f[3]))

    for i, ch in enumerate(chars):
        f = forms.get(ch)
        if not f:
            shaped.append(ch)
            continue

        prev = chars[i - 1] if i > 0 else ""
        nxt = chars[i + 1] if i + 1 < len(chars) else ""

        join_prev = can_join_prev(ch) and can_join_next(prev)
        join_next = can_join_next(ch) and can_join_prev(nxt)

        isolated, final, initial, medial = f
        if join_prev and join_next and medial:
            shaped.append(medial)
        elif join_prev and final:
            shaped.append(final)
        elif join_next and initial:
            shaped.append(initial)
        else:
            shaped.append(isolated)

    return "".join(shaped)


def get_font(size: int, font_path: Optional[str] = None) -> ImageFont.FreeTypeFont | ImageFont.ImageFont:
    candidates = []
    if font_path:
        candidates.append(font_path)
    candidates.extend([
        "C:/Windows/Fonts/segoeui.ttf",
        "C:/Windows/Fonts/tahoma.ttf",
        "C:/Windows/Fonts/arial.ttf",
        "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
        "/usr/share/fonts/truetype/liberation2/LiberationSans-Regular.ttf",
    ])
    for p in candidates:
        try:
            if p and Path(p).exists():
                return ImageFont.truetype(str(p), size=size)
        except Exception:
            pass
    try:
        return ImageFont.truetype("arial.ttf", size=size)
    except Exception:
        return ImageFont.load_default()


def build_qr_image(
    data: str,
    *,
    box_size: int = 10,
    border: int = 4,
    error_level: str = "H",
    fill_color: str = "black",
    back_color: str = "white",
) -> Image.Image:
    qr = qrcode.QRCode(
        version=None,
        error_correction=ERROR_LEVELS[error_level],
        box_size=box_size,
        border=border,
    )
    qr.add_data(data)
    qr.make(fit=True)
    img = qr.make_image(fill_color=fill_color, back_color=back_color)
    if hasattr(img, "get_image"):
        img = img.get_image()
    return img.convert("RGBA")


def add_center_logo(qr_img: Image.Image, logo_path: str, logo_scale_percent: int) -> Image.Image:
    logo_file = Path(logo_path).expanduser()
    if not logo_file.exists():
        raise FileNotFoundError(f"فایل لوگو پیدا نشد: {logo_file}")

    result = qr_img.copy()
    qr_w, qr_h = result.size
    target = max(24, int(min(qr_w, qr_h) * max(5, min(30, logo_scale_percent)) / 100))

    logo = Image.open(logo_file).convert("RGBA")
    logo.thumbnail((target, target), Image.LANCZOS)

    padding = max(8, target // 8)
    bg_w = logo.width + padding * 2
    bg_h = logo.height + padding * 2

    white_bg = Image.new("RGBA", (bg_w, bg_h), (255, 255, 255, 0))
    draw = ImageDraw.Draw(white_bg)
    radius = max(10, min(bg_w, bg_h) // 6)
    draw.rounded_rectangle((0, 0, bg_w - 1, bg_h - 1), radius=radius, fill=(255, 255, 255, 245), outline=(0, 0, 0, 25), width=1)

    bg_x = (result.width - bg_w) // 2
    bg_y = (result.height - bg_h) // 2
    logo_x = bg_x + (bg_w - logo.width) // 2
    logo_y = bg_y + (bg_h - logo.height) // 2

    result.alpha_composite(white_bg, (bg_x, bg_y))
    result.alpha_composite(logo, (logo_x, logo_y))
    return result


def add_frame_and_caption(img: Image.Image, *, caption: str, options: RenderOptions) -> Image.Image:
    img = img.convert("RGBA")
    frame_width = max(0, int(options.frame_width)) if options.show_frame else 0

    caption_lines = []
    if options.show_caption:
        raw_lines = [line.strip() for line in (caption or "").splitlines() if line.strip()]
        caption_lines = [prepare_caption_text(line, options.caption_direction) for line in raw_lines]
    base_font_size = max(10, int(options.caption_font_size or 28))

    text_pad_top = max(12, img.width // 32)
    text_pad_bottom = max(12, img.width // 28)
    text_line_gap = max(4, base_font_size // 5)

    caption_h = 0
    font = get_font(base_font_size, options.caption_font_path)
    caption_bboxes: list[tuple[int, int, int, int]] = []

    canvas_w = img.width + frame_width * 2
    side_pad = max(16, frame_width + 16)
    max_text_width = max(60, canvas_w - side_pad * 2)

    if caption_lines:
        font_size = base_font_size
        while font_size > 10:
            trial_font = get_font(font_size, options.caption_font_path)
            tmp = Image.new("RGBA", (10, 10), (255, 255, 255, 255))
            d = ImageDraw.Draw(tmp)
            trial_bboxes = [d.textbbox((0, 0), line, font=trial_font) for line in caption_lines]
            trial_width = max((box[2] - box[0]) for box in trial_bboxes)
            if trial_width <= max_text_width:
                font = trial_font
                caption_bboxes = trial_bboxes
                break
            font_size -= 1

        if not caption_bboxes:
            tmp = Image.new("RGBA", (10, 10), (255, 255, 255, 255))
            d = ImageDraw.Draw(tmp)
            caption_bboxes = [d.textbbox((0, 0), line, font=font) for line in caption_lines]

        lines_h = sum((box[3] - box[1]) for box in caption_bboxes)
        gaps_h = text_line_gap * max(0, len(caption_bboxes) - 1)
        caption_h = lines_h + gaps_h + text_pad_top + text_pad_bottom

    canvas_h = img.height + frame_width * 2 + caption_h

    canvas = Image.new("RGBA", (canvas_w, canvas_h), (255, 255, 255, 255))
    canvas.alpha_composite(img, (frame_width, frame_width))

    if frame_width > 0:
        draw = ImageDraw.Draw(canvas)
        draw.rectangle((0, 0, canvas_w - 1, canvas_h - 1), outline=(0, 0, 0, 255), width=frame_width)

    if caption_lines and caption_bboxes:
        draw = ImageDraw.Draw(canvas)
        fill = options.caption_color or "#000000"
        text_y = img.height + frame_width + text_pad_top
        for line, box in zip(caption_lines, caption_bboxes):
            text_w = box[2] - box[0]
            text_h = box[3] - box[1]
            text_x = max(0, (canvas_w - text_w) // 2)
            draw.text((text_x, text_y), line, fill=fill, font=font)
            text_y += text_h + text_line_gap

    return canvas


def build_final_image(job: QRJob, options: RenderOptions) -> Image.Image:
    qr_img = build_qr_image(
        job.url,
        box_size=options.box_size,
        border=options.border,
        error_level=options.error_level,
        fill_color=options.fill_color,
        back_color=options.back_color,
    )

    if options.logo_path:
        qr_img = add_center_logo(qr_img, options.logo_path, options.logo_scale_percent)

    caption_text = f"{(job.code or '').strip()}\n{job.name}".strip()
    return add_frame_and_caption(qr_img, caption=caption_text, options=options)


def save_qr_png(job: QRJob, output_dir: Path, options: RenderOptions) -> Path:
    output_dir.mkdir(parents=True, exist_ok=True)
    filename = sanitize_filename(job.name or "qr_code") + ".png"
    out_path = output_dir / filename
    image = build_final_image(job, options)
    image.save(out_path, format="PNG")
    return out_path


def fetch_json_from_url(url: str, timeout: int = 20) -> object:
    req = Request(url, headers={"User-Agent": "Mozilla/5.0 QR Product Generator"})
    with urlopen(req, timeout=timeout) as resp:
        raw = resp.read()
    return json.loads(raw.decode("utf-8-sig"))


def fetch_json_from_file(file_path: str) -> object:
    path = Path(file_path).expanduser()
    if not path.exists():
        raise FileNotFoundError(f"فایل JSON پیدا نشد: {path}")
    return json.loads(path.read_text(encoding="utf-8-sig"))


def import_jobs_from_json_source(
    *,
    source_url: str = "",
    source_file: str = "",
    collection_path: str = "data.products.data",
    title_field: str = "title",
    link_field: str = "id",
    link_template: str = "{id}",
) -> List[QRJob]:
    source_url = (source_url or "").strip()
    source_file = (source_file or "").strip()

    if source_file:
        payload = fetch_json_from_file(source_file)
    elif source_url:
        payload = fetch_json_from_url(source_url)
    else:
        raise ValueError("اول لینک JSON یا فایل JSON را وارد کنید.")

    items = read_nested_value(payload, collection_path)

    if not isinstance(items, list):
        raise ValueError("مسیر داده باید به یک لیست برسد.")

    jobs: List[QRJob] = []
    for i, item in enumerate(items, start=1):
        if not isinstance(item, dict):
            continue

        raw_name = item.get(title_field, f"product_{i}")
        raw_link_value = item.get(link_field, "")

        if raw_link_value is None or raw_link_value == "":
            continue

        name = str(raw_name).strip() or f"product_{i}"
        raw_link_text = str(raw_link_value).strip()

        if "{" in link_template and "}" in link_template:
            try:
                url = link_template.format(**item)
            except Exception:
                url = link_template.replace("{id}", raw_link_text)
        else:
            if is_valid_url(raw_link_text):
                url = raw_link_text
            else:
                base = (link_template or "").strip()
                if not base or base == "{id}":
                    url = raw_link_text
                else:
                    url = f"{base.rstrip('/')}/{raw_link_text}"

        url = str(url).strip()
        if not is_valid_url(url):
            continue

        jobs.append(QRJob(name=name, url=url))

    if not jobs:
        raise ValueError("هیچ آیتم معتبری از ورودی JSON استخراج نشد. مسیر داده یا قالب لینک را بررسی کنید.")
    return jobs


def read_nested_value(data: object, path: str) -> object:
    current = data
    clean = (path or "").strip().replace(">", ".")
    parts = [p.strip() for p in clean.split(".") if p.strip()]
    for part in parts:
        if isinstance(current, dict):
            if part not in current:
                raise KeyError(f"مسیر پیدا نشد: {path}")
            current = current[part]
        elif isinstance(current, list):
            if not part.isdigit():
                raise KeyError(f"برای ورود به لیست باید شماره اندیس بدهید: {path}")
            idx = int(part)
            current = current[idx]
        else:
            raise KeyError(f"مسیر پیدا نشد: {path}")
    return current


def import_jobs_from_json_url(
    source_url: str,
    *,
    collection_path: str = "data.products.data",
    title_field: str = "title",
    link_field: str = "id",
    link_template: str = "{id}",
) -> List[QRJob]:
    return import_jobs_from_json_source(
        source_url=source_url,
        collection_path=collection_path,
        title_field=title_field,
        link_field=link_field,
        link_template=link_template,
    )




def extract_first_image_url(page_url: str) -> str:
    req = Request(page_url, headers={"User-Agent": "Mozilla/5.0"})
    with urlopen(req, timeout=20) as resp:
        charset = resp.headers.get_content_charset() or "utf-8"
        body = resp.read().decode(charset, errors="ignore")

    patterns = [
        r'<meta[^>]+property=["\']og:image["\'][^>]+content=["\']([^"\']+)',
        r'<meta[^>]+content=["\']([^"\']+)["\'][^>]+property=["\']og:image["\']',
        r'<img[^>]+src=["\']([^"\']+)["\']',
    ]
    for pat in patterns:
        m = re.search(pat, body, flags=re.IGNORECASE)
        if m:
            src = html.unescape(m.group(1).strip())
            return src if src.startswith("http") else ""
    return ""


def download_product_images(jobs: List[QRJob], output_dir: Path) -> List[Path]:
    output_dir.mkdir(parents=True, exist_ok=True)
    files: List[Path] = []
    for i, job in enumerate(jobs, start=1):
        img_url = extract_first_image_url(job.url)
        if not img_url:
            continue
        req = Request(img_url, headers={"User-Agent": "Mozilla/5.0"})
        with urlopen(req, timeout=30) as resp:
            content_type = (resp.headers.get("Content-Type") or "").lower()
            ext = ".jpg"
            if "png" in content_type:
                ext = ".png"
            elif "webp" in content_type:
                ext = ".webp"
            filename = sanitize_filename(job.name or f"product_{i}") + ext
            path = output_dir / filename
            path.write_bytes(resp.read())
            files.append(path)
    return files


def parse_excel_file(path: Path) -> List[QRJob]:
    try:
        from openpyxl import load_workbook
    except Exception as exc:  # pragma: no cover
        raise ValueError("برای خواندن Excel باید openpyxl نصب باشد: pip install openpyxl") from exc

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError("فایل اکسل خالی است.")
    headers = [str(x).strip().lower() if x is not None else "" for x in rows[0]]
    def find_idx(cands):
        for c in cands:
            if c in headers:
                return headers.index(c)
        return -1
    name_idx = find_idx(["name","title","product","نام"])
    url_idx = find_idx(["url","link","href","لینک"])
    if url_idx < 0:
        raise ValueError("در Excel باید ستون url یا link وجود داشته باشد.")
    jobs=[]
    for i, row in enumerate(rows[1:], start=1):
        vals=list(row)
        url=str(vals[url_idx]).strip() if url_idx < len(vals) and vals[url_idx] is not None else ""
        name=(str(vals[name_idx]).strip() if name_idx >=0 and name_idx < len(vals) and vals[name_idx] is not None else f"product_{i}")
        if not url:
            continue
        if not is_valid_url(url):
            raise ValueError(f"ردیف {i}: لینک نامعتبر است -> {url}")
        jobs.append(QRJob(name=name, url=url))
    if not jobs:
        raise ValueError("هیچ ردیف معتبری پیدا نشد.")
    return jobs

def parse_batch_lines(text: str) -> List[QRJob]:
    jobs: List[QRJob] = []
    for idx, raw_line in enumerate(text.splitlines(), start=1):
        line = raw_line.strip()
        if not line or line.startswith("نمونه") or line.startswith("یا فقط"):
            continue

        if "," in line:
            left, right = line.split(",", 1)
            name = left.strip()
            url = right.strip()
        else:
            name = f"product_{idx}"
            url = line

        if not is_valid_url(url):
            raise ValueError(f"خط {idx}: لینک نامعتبر است -> {url}")

        jobs.append(QRJob(name=name, url=url))

    if not jobs:
        raise ValueError("هیچ لینک معتبری وارد نشده است.")
    return jobs


def parse_batch_file(path: Path) -> List[QRJob]:
    suffix = path.suffix.lower()
    jobs: List[QRJob] = []

    if suffix in {".xlsx", ".xlsm"}:
        jobs = parse_excel_file(path)
    elif suffix in {".csv", ".tsv"}:
        delimiter = "\t" if suffix == ".tsv" else ","
        with path.open("r", encoding="utf-8-sig", newline="") as f:
            reader = csv.DictReader(f, delimiter=delimiter)
            headers = {h.lower().strip(): h for h in (reader.fieldnames or [])}
            name_key = headers.get("name") or headers.get("title") or headers.get("product")
            url_key = headers.get("url") or headers.get("link") or headers.get("href")
            if not url_key:
                raise ValueError("فایل CSV/TSV باید ستون url یا link داشته باشد.")
            for i, row in enumerate(reader, start=1):
                url = (row.get(url_key) or "").strip()
                name = (row.get(name_key) or f"product_{i}").strip() if name_key else f"product_{i}"
                if not is_valid_url(url):
                    raise ValueError(f"ردیف {i}: لینک نامعتبر است -> {url}")
                jobs.append(QRJob(name=name, url=url))
    else:
        jobs = parse_batch_lines(path.read_text(encoding="utf-8-sig"))

    if not jobs:
        raise ValueError("هیچ موردی برای ساخت QR پیدا نشد.")
    return jobs


def zip_pngs(files: Iterable[Path], zip_path: Path) -> Path:
    zip_path.parent.mkdir(parents=True, exist_ok=True)
    with zipfile.ZipFile(zip_path, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        for file in files:
            zf.write(file, arcname=file.name)
    return zip_path


class QRApp:
    def __init__(self, root: tk.Tk):
        self.root = root
        self.root.title("سازنده کیوآرکد محصولات")
        self.root.geometry("1100x860")
        self.root.minsize(980, 740)

        self.preview_photo: Optional[ImageTk.PhotoImage] = None
        self.preview_source_image: Optional[Image.Image] = None

        self.single_url_var = tk.StringVar()
        self.single_code_var = tk.StringVar()
        self.single_name_var = tk.StringVar(value="product_qr")

        self.box_size_var = tk.IntVar(value=10)
        self.border_var = tk.IntVar(value=4)
        self.error_var = tk.StringVar(value="H")
        self.output_dir_var = tk.StringVar(value=str(Path.cwd() / "qr_output"))
        self.batch_file_var = tk.StringVar()
        self.status_var = tk.StringVar(value="آماده")

        self.logo_path_var = tk.StringVar()
        self.logo_scale_var = tk.IntVar(value=18)
        self.show_frame_var = tk.BooleanVar(value=True)
        self.frame_width_var = tk.IntVar(value=12)
        self.show_caption_var = tk.BooleanVar(value=True)
        self.font_path_var = tk.StringVar()
        self.font_size_var = tk.IntVar(value=28)
        self.font_color_var = tk.StringVar(value="#000000")
        self.direction_var = tk.StringVar(value="خودکار")

        self.import_url_var = tk.StringVar()
        self.import_file_var = tk.StringVar()
        self.import_path_var = tk.StringVar(value="data.products.data")
        self.import_title_var = tk.StringVar(value="title")
        self.import_link_var = tk.StringVar(value="id")
        self.import_link_template_var = tk.StringVar(value="{id}")

        self._build_ui()

    def _build_ui(self) -> None:
        pad = {"padx": 10, "pady": 8}

        top = ttk.Frame(self.root)
        top.pack(fill="x", padx=12, pady=10)

        options = ttk.LabelFrame(top, text="تنظیمات خروجی تصویر")
        options.pack(fill="x")

        ttk.Label(options, text="کیفیت / اندازه ماژول").grid(row=0, column=0, sticky="w", **pad)
        ttk.Spinbox(options, from_=4, to=40, textvariable=self.box_size_var, width=8).grid(row=0, column=1, sticky="w", **pad)

        ttk.Label(options, text="حاشیه داخلی کیوآرکد").grid(row=0, column=2, sticky="w", **pad)
        ttk.Spinbox(options, from_=1, to=20, textvariable=self.border_var, width=8).grid(row=0, column=3, sticky="w", **pad)

        ttk.Label(options, text="میزان تصحیح خطا").grid(row=0, column=4, sticky="w", **pad)
        ttk.Combobox(options, values=["L", "M", "Q", "H"], textvariable=self.error_var, width=6, state="readonly").grid(row=0, column=5, sticky="w", **pad)

        ttk.Label(options, text="لوگوی وسط کیوآرکد").grid(row=1, column=0, sticky="w", **pad)
        ttk.Entry(options, textvariable=self.logo_path_var, justify="right").grid(row=1, column=1, columnspan=3, sticky="ew", **pad)
        ttk.Button(options, text="انتخاب لوگو", command=self.choose_logo_file).grid(row=1, column=4, sticky="ew", **pad)
        ttk.Button(options, text="پاک کردن لوگو", command=lambda: self.logo_path_var.set("")).grid(row=1, column=5, sticky="ew", **pad)

        ttk.Label(options, text="اندازه لوگو %").grid(row=2, column=0, sticky="w", **pad)
        ttk.Spinbox(options, from_=5, to=30, textvariable=self.logo_scale_var, width=8).grid(row=2, column=1, sticky="w", **pad)

        ttk.Checkbutton(options, text="کادر دور تصویر", variable=self.show_frame_var, command=self.safe_preview).grid(row=2, column=2, sticky="w", **pad)
        ttk.Label(options, text="ضخامت کادر").grid(row=2, column=3, sticky="w", **pad)
        ttk.Spinbox(options, from_=0, to=40, textvariable=self.frame_width_var, width=8).grid(row=2, column=4, sticky="w", **pad)

        ttk.Checkbutton(options, text="نمایش نام زیر کیوآرکد", variable=self.show_caption_var, command=self.safe_preview).grid(row=3, column=0, sticky="w", **pad)
        ttk.Label(options, text="اندازه فونت").grid(row=3, column=1, sticky="w", **pad)
        ttk.Spinbox(options, from_=8, to=96, textvariable=self.font_size_var, width=8).grid(row=3, column=2, sticky="w", **pad)
        ttk.Label(options, text="رنگ متن").grid(row=3, column=3, sticky="w", **pad)
        color_row = ttk.Frame(options)
        color_row.grid(row=3, column=4, columnspan=2, sticky="ew", **pad)
        color_row.columnconfigure(0, weight=1)
        ttk.Entry(color_row, textvariable=self.font_color_var, width=12, justify="right").grid(row=0, column=0, sticky="ew")
        ttk.Button(color_row, text="انتخاب رنگ", command=self.choose_font_color).grid(row=0, column=1, padx=(8, 0))

        ttk.Label(options, text="جهت نوشتار").grid(row=4, column=0, sticky="w", **pad)
        ttk.Combobox(options, values=["خودکار", "راست‌به‌چپ", "چپ‌به‌راست"], textvariable=self.direction_var, width=10, state="readonly").grid(row=4, column=1, sticky="w", **pad)
        ttk.Label(options, text="خودکار = فارسی‌ها راست‌به‌چپ، انگلیسی‌ها چپ‌به‌راست").grid(row=4, column=2, columnspan=4, sticky="w", **pad)

        ttk.Label(options, text="فونت متن").grid(row=5, column=0, sticky="w", **pad)
        ttk.Entry(options, textvariable=self.font_path_var, justify="right").grid(row=5, column=1, columnspan=3, sticky="ew", **pad)
        ttk.Button(options, text="انتخاب فونت", command=self.choose_font_file).grid(row=5, column=4, sticky="ew", **pad)
        ttk.Button(options, text="فونت پیش‌فرض", command=lambda: self.font_path_var.set("")).grid(row=5, column=5, sticky="ew", **pad)

        ttk.Label(options, text="پوشه خروجی").grid(row=6, column=0, sticky="w", **pad)
        ttk.Entry(options, textvariable=self.output_dir_var, justify="right").grid(row=6, column=1, columnspan=4, sticky="ew", **pad)
        ttk.Button(options, text="انتخاب پوشه", command=self.choose_output_dir).grid(row=6, column=5, sticky="ew", **pad)

        for idx in range(6):
            options.columnconfigure(idx, weight=1 if idx in {1, 2, 3, 4} else 0)

        body = ttk.Frame(self.root)
        body.pack(fill="both", expand=True, padx=12, pady=(0, 12))
        body.columnconfigure(0, weight=3)
        body.columnconfigure(1, weight=2)
        body.rowconfigure(0, weight=1)

        left = ttk.Notebook(body)
        left.grid(row=0, column=0, sticky="nsew", padx=(0, 10))

        single_tab = ttk.Frame(left)
        batch_tab = ttk.Frame(left)
        left.add(single_tab, text="تک‌لینک")
        left.add(batch_tab, text="چند‌لینک")

        self._build_single_tab(single_tab)
        self._build_batch_tab(batch_tab)

        self.preview_frame = ttk.LabelFrame(body, text="پیش‌نمایش")
        self.preview_frame.grid(row=0, column=1, sticky="nsew")
        self.preview_frame.rowconfigure(0, weight=1)
        self.preview_frame.columnconfigure(0, weight=1)

        self.preview_label = ttk.Label(self.preview_frame, text="برای دیدن پیش‌نمایش، یک لینک وارد کنید.", anchor="center", justify="center")
        self.preview_label.grid(row=0, column=0, sticky="nsew", padx=12, pady=12)
        self.preview_frame.bind("<Configure>", self._on_preview_resize)
        self.preview_label.bind("<Configure>", self._on_preview_resize)

        status = ttk.Label(self.root, textvariable=self.status_var, anchor="w")
        status.pack(fill="x", padx=12, pady=(0, 8))

    def _build_single_tab(self, parent: ttk.Frame) -> None:
        pad = {"padx": 10, "pady": 8}
        frame = ttk.LabelFrame(parent, text="ساخت تصویر برای یک محصول")
        frame.pack(fill="both", expand=True, padx=10, pady=10)
        frame.columnconfigure(1, weight=1)

        ttk.Label(frame, text="کد محصول").grid(row=0, column=0, sticky="w", **pad)
        code_row = ttk.Frame(frame)
        code_row.grid(row=0, column=1, sticky="ew", **pad)
        code_row.columnconfigure(0, weight=1)
        code_entry = ttk.Entry(code_row, textvariable=self.single_code_var)
        code_entry.grid(row=0, column=0, sticky="ew")
        ttk.Button(code_row, text="چسباندن", command=lambda: self.paste_to_stringvar(self.single_code_var, preview_after=True)).grid(row=0, column=1, padx=(8, 0))
        code_entry.bind("<Control-v>", lambda _e: (self.paste_to_stringvar(self.single_code_var, preview_after=True), "break")[1])
        code_entry.bind("<Shift-Insert>", lambda _e: (self.paste_to_stringvar(self.single_code_var, preview_after=True), "break")[1])
        code_entry.bind("<KeyRelease>", lambda _e: self.safe_preview())

        ttk.Label(frame, text="نام فایل / متن زیر کیوآرکد").grid(row=1, column=0, sticky="w", **pad)
        name_row = ttk.Frame(frame)
        name_row.grid(row=1, column=1, sticky="ew", **pad)
        name_row.columnconfigure(0, weight=1)
        name_entry = ttk.Entry(name_row, textvariable=self.single_name_var)
        name_entry.grid(row=0, column=0, sticky="ew")
        ttk.Button(name_row, text="چسباندن", command=lambda: self.paste_to_stringvar(self.single_name_var, preview_after=True)).grid(row=0, column=1, padx=(8, 0))
        name_entry.bind("<Control-v>", lambda _e: (self.paste_to_stringvar(self.single_name_var, preview_after=True), "break")[1])
        name_entry.bind("<Shift-Insert>", lambda _e: (self.paste_to_stringvar(self.single_name_var, preview_after=True), "break")[1])
        name_entry.bind("<KeyRelease>", lambda _e: self.safe_preview())

        ttk.Label(frame, text="لینک محصول").grid(row=2, column=0, sticky="nw", **pad)
        url_row = ttk.Frame(frame)
        url_row.grid(row=2, column=1, sticky="ew", **pad)
        url_row.columnconfigure(0, weight=1)
        url_entry = ttk.Entry(url_row, textvariable=self.single_url_var, justify="left")
        url_entry.grid(row=0, column=0, sticky="ew")
        ttk.Button(url_row, text="چسباندن لینک", command=lambda: self.paste_to_stringvar(self.single_url_var, preview_after=True)).grid(row=0, column=1, padx=(8, 0))
        url_entry.bind("<KeyRelease>", lambda _e: self.safe_preview())
        url_entry.bind("<Control-v>", lambda _e: (self.paste_to_stringvar(self.single_url_var, preview_after=True), "break")[1])
        url_entry.bind("<Shift-Insert>", lambda _e: (self.paste_to_stringvar(self.single_url_var, preview_after=True), "break")[1])

        actions = ttk.Frame(frame)
        actions.grid(row=3, column=0, columnspan=2, sticky="w", **pad)
        ttk.Button(actions, text="پیش‌نمایش", command=self.preview_single).pack(side="left", padx=(0, 8))
        ttk.Button(actions, text="ذخیره تصویر", command=self.save_single).pack(side="left")

        help_text = (
            "می‌توانید فونت فارسی با فایل‌های TTF یا OTF اضافه کنید.\n"
            "رنگ و اندازه متن زیر کیوآرکد قابل تغییر است.\n"
            "فارسی‌ها خودکار راست‌به‌چپ و انگلیسی‌ها خودکار چپ‌به‌راست نمایش داده می‌شوند."
        )
        ttk.Label(frame, text=help_text, justify="left").grid(row=4, column=0, columnspan=2, sticky="w", **pad)

    def _build_batch_tab(self, parent: ttk.Frame) -> None:
        pad = {"padx": 10, "pady": 8}

        frame = ttk.LabelFrame(parent, text="ساخت گروهی تصویر")
        frame.pack(fill="both", expand=True, padx=10, pady=10)
        frame.columnconfigure(1, weight=1)
        frame.rowconfigure(2, weight=1)

        ttk.Label(frame, text="فایل CSV یا TXT").grid(row=0, column=0, sticky="w", **pad)
        ttk.Entry(frame, textvariable=self.batch_file_var, justify="right").grid(row=0, column=1, sticky="ew", **pad)
        ttk.Button(frame, text="انتخاب فایل", command=self.choose_batch_file).grid(row=0, column=2, sticky="ew", **pad)

        batch_header = ttk.Frame(frame)
        batch_header.grid(row=1, column=0, columnspan=3, sticky="ew", padx=10, pady=(8, 0))
        batch_header.columnconfigure(0, weight=1)
        ttk.Label(batch_header, text="یا فهرست را اینجا بچسبانید").grid(row=0, column=0, sticky="w")
        ttk.Button(batch_header, text="چسباندن از کلیپ‌بورد", command=lambda: self.paste_to_text(self.batch_text)).grid(row=0, column=1, sticky="e")

        self.batch_text = ScrolledText(frame, wrap="word", height=16)
        self.batch_text.grid(row=2, column=0, columnspan=3, sticky="nsew", padx=10, pady=(8, 10))
        self.batch_text.bind("<Control-v>", lambda _e: (self.paste_to_text(self.batch_text), "break")[1])
        self.batch_text.bind("<Shift-Insert>", lambda _e: (self.paste_to_text(self.batch_text), "break")[1])
        self.batch_text.insert(
            "1.0",
            "نمونه:\n"
            "product-1,https://example.com/products/1\n"
            "product-2,https://example.com/products/2\n\n"
            "یا فقط لینک:\n"
            "https://example.com/products/3\n",
        )

        import_box = ttk.LabelFrame(frame, text="ورود خودکار از JSON")
        import_box.grid(row=3, column=0, columnspan=3, sticky="ew", padx=10, pady=(0, 10))
        for c in range(5):
            import_box.columnconfigure(c, weight=1 if c in {1, 3} else 0)

        ttk.Label(import_box, text="لینک منبع").grid(row=0, column=0, sticky="w", padx=8, pady=6)
        ttk.Entry(import_box, textvariable=self.import_url_var, justify="right").grid(row=0, column=1, columnspan=4, sticky="ew", padx=8, pady=6)

        ttk.Label(import_box, text="فایل JSON").grid(row=1, column=0, sticky="w", padx=8, pady=6)
        ttk.Entry(import_box, textvariable=self.import_file_var, justify="right").grid(row=1, column=1, columnspan=3, sticky="ew", padx=8, pady=6)
        ttk.Button(import_box, text="انتخاب فایل JSON", command=self.choose_json_import_file).grid(row=1, column=4, sticky="ew", padx=8, pady=6)

        ttk.Label(import_box, text="مسیر داده").grid(row=2, column=0, sticky="w", padx=8, pady=6)
        ttk.Entry(import_box, textvariable=self.import_path_var, justify="right").grid(row=2, column=1, sticky="ew", padx=8, pady=6)
        ttk.Label(import_box, text="فیلد نام").grid(row=2, column=2, sticky="w", padx=8, pady=6)
        ttk.Entry(import_box, textvariable=self.import_title_var, justify="right", width=18).grid(row=2, column=3, sticky="ew", padx=8, pady=6)
        ttk.Label(import_box, text="فیلد لینک").grid(row=2, column=4, sticky="w", padx=8, pady=6)

        ttk.Entry(import_box, textvariable=self.import_link_var, justify="right").grid(row=3, column=1, sticky="ew", padx=8, pady=6)
        ttk.Label(import_box, text="قالب لینک").grid(row=3, column=2, sticky="w", padx=8, pady=6)
        ttk.Entry(import_box, textvariable=self.import_link_template_var, justify="right").grid(row=3, column=3, columnspan=2, sticky="ew", padx=8, pady=6)

        ttk.Label(import_box, text="می‌توانی یا لینک بدهی یا فایل JSON انتخاب کنی. قالب لینک می‌تواند {id} یا آدرس پایه مثل https://site.com/product باشد.").grid(row=4, column=0, columnspan=5, sticky="w", padx=8, pady=(0, 6))

        import_actions = ttk.Frame(import_box)
        import_actions.grid(row=5, column=0, columnspan=5, sticky="w", padx=8, pady=(0, 8))
        ttk.Button(import_actions, text="افزودن به لیست", command=self.import_json_append).pack(side="left", padx=(0, 8))
        ttk.Button(import_actions, text="جایگزینی لیست", command=self.import_json_replace).pack(side="left")

        actions = ttk.Frame(frame)
        actions.grid(row=4, column=0, columnspan=3, sticky="w", **pad)
        ttk.Button(actions, text="ساخت همه تصویرها", command=self.generate_batch).pack(side="left", padx=(0, 8))
        ttk.Button(actions, text="ساخت تصویرها + ZIP", command=self.generate_batch_zip).pack(side="left")

        help_text = (
            "فرمت پیشنهادی CSV: ستون‌های name و url\n"
            "فرمت متن ساده: هر خط به صورت name,url یا فقط url\n"
            "فونت، رنگ متن، لوگو، کادر و نام زیر کیوآرکد روی همه خروجی‌ها اعمال می‌شود. همچنین می‌توانید داده‌ها را مستقیم از لینک JSON یا فایل JSON وارد لیست کنید."
        )
        ttk.Label(frame, text=help_text, justify="left").grid(row=5, column=0, columnspan=3, sticky="w", **pad)

    def paste_to_stringvar(self, var: tk.StringVar, *, append: bool = False, preview_after: bool = False) -> None:
        try:
            value = self.root.clipboard_get()
        except Exception:
            self.status_var.set("کلیپ‌بورد خالی است یا خوانده نشد.")
            return
        if not append:
            var.set(value.strip())
        else:
            current = var.get().strip()
            var.set((current + value).strip())
        if preview_after:
            self.safe_preview()
        self.status_var.set("متن از کلیپ‌بورد چسبانده شد.")

    def paste_to_text(self, widget: tk.Text, *, replace: bool = False) -> None:
        try:
            value = self.root.clipboard_get()
        except Exception:
            self.status_var.set("کلیپ‌بورد خالی است یا خوانده نشد.")
            return
        if replace:
            widget.delete("1.0", "end")
        widget.insert(widget.index("insert"), value)
        self.status_var.set("متن از کلیپ‌بورد چسبانده شد.")

    def _on_preview_resize(self, _event=None) -> None:
        if self.preview_source_image is not None:
            self._render_preview_to_fit()

    def _get_preview_target_size(self) -> tuple[int, int]:
        self.root.update_idletasks()
        target_w = max(120, self.preview_label.winfo_width() - 8)
        target_h = max(120, self.preview_label.winfo_height() - 8)

        if target_w <= 120 or target_h <= 120:
            try:
                target_w = max(120, self.preview_frame.winfo_width() - 32)
                target_h = max(120, self.preview_frame.winfo_height() - 32)
            except Exception:
                target_w, target_h = 420, 560

        return target_w, target_h

    def _render_preview_to_fit(self) -> None:
        if self.preview_source_image is None:
            return
        target_w, target_h = self._get_preview_target_size()
        preview = self.preview_source_image.copy()
        preview.thumbnail((target_w, target_h), Image.LANCZOS)
        self.preview_photo = ImageTk.PhotoImage(preview)
        self.preview_label.configure(image=self.preview_photo, text="")

    def choose_output_dir(self) -> None:
        folder = filedialog.askdirectory(initialdir=self.output_dir_var.get() or str(Path.cwd()))
        if folder:
            self.output_dir_var.set(folder)

    def choose_batch_file(self) -> None:
        path = filedialog.askopenfilename(
            filetypes=[("CSV files", "*.csv"), ("TSV files", "*.tsv"), ("Text files", "*.txt"), ("All files", "*.*")]
        )
        if path:
            self.batch_file_var.set(path)

    def choose_json_import_file(self) -> None:
        path = filedialog.askopenfilename(
            filetypes=[("JSON files", "*.json"), ("All files", "*.*")]
        )
        if path:
            self.import_file_var.set(path)

    def choose_logo_file(self) -> None:
        path = filedialog.askopenfilename(
            filetypes=[("Image files", "*.png;*.jpg;*.jpeg;*.webp;*.bmp"), ("All files", "*.*")]
        )
        if path:
            self.logo_path_var.set(path)
            self.safe_preview()

    def choose_font_file(self) -> None:
        path = filedialog.askopenfilename(
            filetypes=[("Font files", "*.ttf;*.otf;*.ttc"), ("All files", "*.*")]
        )
        if path:
            self.font_path_var.set(path)
            self.safe_preview()

    def choose_font_color(self) -> None:
        if colorchooser is None:
            return
        color = colorchooser.askcolor(title="انتخاب رنگ متن", color=self.font_color_var.get())
        if color and color[1]:
            self.font_color_var.set(color[1])
            self.safe_preview()

    def current_options(self) -> RenderOptions:
        error_level = self.error_var.get().strip().upper() or "H"
        if error_level not in ERROR_LEVELS:
            error_level = "H"
        return RenderOptions(
            box_size=max(4, int(self.box_size_var.get())),
            border=max(1, int(self.border_var.get())),
            error_level=error_level,
            logo_path=self.logo_path_var.get().strip() or None,
            logo_scale_percent=max(5, min(30, int(self.logo_scale_var.get()))),
            show_frame=bool(self.show_frame_var.get()),
            frame_width=max(0, int(self.frame_width_var.get())),
            show_caption=bool(self.show_caption_var.get()),
            caption_font_path=self.font_path_var.get().strip() or None,
            caption_font_size=max(8, min(200, int(self.font_size_var.get()))),
            caption_color=(self.font_color_var.get().strip() or "#000000"),
            caption_direction=normalize_direction_value(self.direction_var.get()),
        )

    def safe_preview(self) -> None:
        url = self.single_url_var.get().strip()
        if is_valid_url(url):
            try:
                self.preview_single()
            except Exception:
                pass

    def preview_single(self) -> None:
        url = self.single_url_var.get().strip()
        if not is_valid_url(url):
            self.status_var.set("لینک محصول معتبر نیست.")
            return

        name = self.single_name_var.get().strip() or "product_qr"
        code = self.single_code_var.get().strip()
        self.preview_source_image = build_final_image(QRJob(name=name, url=url, code=code), self.current_options())
        self._render_preview_to_fit()
        self.status_var.set("پیش‌نمایش به‌روزرسانی شد و با اندازه پنجره تنظیم شد.")

    def save_single(self) -> None:
        url = self.single_url_var.get().strip()
        name = self.single_name_var.get().strip() or "product_qr"
        code = self.single_code_var.get().strip()
        if not is_valid_url(url):
            messagebox.showerror("خطا", "لینک محصول معتبر نیست.")
            return

        out_dir = Path(self.output_dir_var.get()).expanduser()
        path = save_qr_png(QRJob(name=name, url=url, code=code), out_dir, self.current_options())
        self.status_var.set(f"فایل ذخیره شد: {path}")
        messagebox.showinfo("انجام شد", f"تصویر ذخیره شد:\n{path}")

    def _collect_batch_jobs(self) -> List[QRJob]:
        file_path = self.batch_file_var.get().strip()
        pasted = self.batch_text.get("1.0", "end").strip()
        if file_path:
            return parse_batch_file(Path(file_path))
        return parse_batch_lines(pasted)

    def _import_jobs_from_current_source(self) -> List[QRJob]:
        source_url = self.import_url_var.get().strip()
        source_file = self.import_file_var.get().strip()

        if source_url and not is_valid_url(source_url):
            raise ValueError("لینک منبع معتبر نیست.")
        if not source_url and not source_file:
            raise ValueError("اول لینک JSON یا فایل JSON را وارد کنید.")

        collection_path = self.import_path_var.get().strip() or "data.products.data"
        title_field = self.import_title_var.get().strip() or "title"
        link_field = self.import_link_var.get().strip() or "id"
        link_template = self.import_link_template_var.get().strip() or "{id}"

        return import_jobs_from_json_source(
            source_url=source_url,
            source_file=source_file,
            collection_path=collection_path,
            title_field=title_field,
            link_field=link_field,
            link_template=link_template,
        )

    def _write_jobs_to_batch_text(self, jobs: List[QRJob], *, replace: bool) -> None:
        lines = [f"{job.name},{job.url}" for job in jobs]
        payload = "\n".join(lines) + "\n"
        if replace:
            self.batch_text.delete("1.0", "end")
            self.batch_text.insert("1.0", payload)
        else:
            existing = self.batch_text.get("1.0", "end").strip()
            if existing:
                self.batch_text.insert("end", ("" if existing.endswith("\n") else "\n") + payload)
            else:
                self.batch_text.insert("1.0", payload)

    def import_json_append(self) -> None:
        try:
            jobs = self._import_jobs_from_current_source()
            self._write_jobs_to_batch_text(jobs, replace=False)
            self.status_var.set(f"{len(jobs)} آیتم از JSON خوانده شد و به لیست اضافه شد.")
            messagebox.showinfo("انجام شد", f"{len(jobs)} آیتم به لیست اضافه شد.")
        except Exception as e:
            messagebox.showerror("خطا", str(e))
            self.status_var.set(f"خطا: {e}")

    def import_json_replace(self) -> None:
        try:
            jobs = self._import_jobs_from_current_source()
            self._write_jobs_to_batch_text(jobs, replace=True)
            self.status_var.set(f"{len(jobs)} آیتم از JSON خوانده شد و جایگزین لیست شد.")
            messagebox.showinfo("انجام شد", f"{len(jobs)} آیتم جایگزین لیست شد.")
        except Exception as e:
            messagebox.showerror("خطا", str(e))
            self.status_var.set(f"خطا: {e}")

    def generate_batch(self) -> None:
        try:
            jobs = self._collect_batch_jobs()
            out_dir = Path(self.output_dir_var.get()).expanduser()
            options = self.current_options()
            files = [save_qr_png(job, out_dir, options) for job in jobs]
            self.status_var.set(f"{len(files)} تصویر ساخته شد.")
            messagebox.showinfo("انجام شد", f"{len(files)} تصویر ساخته شد در:\n{out_dir}")
        except Exception as e:
            messagebox.showerror("خطا", str(e))
            self.status_var.set(f"خطا: {e}")

    def generate_batch_zip(self) -> None:
        try:
            jobs = self._collect_batch_jobs()
            out_dir = Path(self.output_dir_var.get()).expanduser()
            options = self.current_options()
            files = [save_qr_png(job, out_dir, options) for job in jobs]
            zip_path = zip_pngs(files, out_dir / "qr_codes_bundle.zip")
            self.status_var.set(f"{len(files)} تصویر ساخته شد و ZIP ایجاد شد.")
            messagebox.showinfo("انجام شد", f"ZIP ساخته شد:\n{zip_path}")
        except Exception as e:
            messagebox.showerror("خطا", str(e))
            self.status_var.set(f"خطا: {e}")


def run_cli(args: argparse.Namespace) -> int:
    error_level = args.error_correction.upper()
    if error_level not in ERROR_LEVELS:
        raise SystemExit("error_correction must be one of: L, M, Q, H")

    options = RenderOptions(
        box_size=max(4, args.box_size),
        border=max(1, args.border),
        error_level=error_level,
        logo_path=args.logo,
        logo_scale_percent=max(5, min(30, args.logo_scale)),
        show_frame=not args.no_frame,
        frame_width=max(0, args.frame_width),
        show_caption=not args.no_caption,
        caption_font_path=args.font,
        caption_font_size=max(8, args.font_size),
        caption_color=args.font_color or "#000000",
        caption_direction=normalize_direction_value(args.direction),
    )
    output_dir = Path(args.output_dir).expanduser()

    if args.url:
        if not is_valid_url(args.url):
            raise SystemExit("URL is not valid. It must start with http:// or https://")
        name = args.name or "product_qr"
        path = save_qr_png(QRJob(name=name, url=args.url), output_dir, options)
        print(path)
        return 0

    if args.batch_file:
        jobs = parse_batch_file(Path(args.batch_file).expanduser())
        if args.download_images:
            files = download_product_images(jobs, output_dir)
            for file in files:
                print(file)
            return 0
        files = [save_qr_png(job, output_dir, options) for job in jobs]
        if args.zip:
            zip_path = zip_pngs(files, output_dir / "qr_codes_bundle.zip")
            print(zip_path)
        else:
            for file in files:
                print(file)
        return 0

    raise SystemExit("Use --url for one QR or --batch-file for batch mode. Without arguments the GUI is launched.")


def build_arg_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="ساخت کیوآرکد محصولات به‌صورت تصویر")
    parser.add_argument("--url", help="لینک یک محصول")
    parser.add_argument("--name", help="نام فایل خروجی بدون پسوند")
    parser.add_argument("--batch-file", help="فایل ورودی CSV/TSV/TXT/Excel برای ساخت گروهی")
    parser.add_argument("--output-dir", default="qr_output", help="پوشه خروجی تصویرها")
    parser.add_argument("--box-size", type=int, default=10, help="اندازه ماژول‌های کیوآرکد")
    parser.add_argument("--border", type=int, default=4, help="حاشیه داخلی کیوآرکد")
    parser.add_argument("--error-correction", default="H", choices=["L", "M", "Q", "H"], help="میزان تصحیح خطا")
    parser.add_argument("--logo", help="مسیر لوگوی وسط کیوآرکد")
    parser.add_argument("--logo-scale", type=int, default=18, help="اندازه لوگو به درصد عرض کیوآرکد")
    parser.add_argument("--frame-width", type=int, default=12, help="ضخامت کادر دور تصویر نهایی")
    parser.add_argument("--no-frame", action="store_true", help="خاموش کردن کادر دور تصویر")
    parser.add_argument("--font", help="مسیر فایل فونت برای متن زیر کیوآرکد")
    parser.add_argument("--font-size", type=int, default=28, help="اندازه فونت متن زیر کیوآرکد")
    parser.add_argument("--font-color", default="#000000", help="رنگ متن، مثل #ff0000")
    parser.add_argument("--direction", default="auto", choices=["auto", "rtl", "ltr"], help="جهت نوشتار متن")
    parser.add_argument("--no-caption", action="store_true", help="خاموش کردن متن زیر کیوآرکد")
    parser.add_argument("--zip", action="store_true", help="ساخت فایل ZIP در حالت گروهی")
    parser.add_argument("--download-images", action="store_true", help="در حالت گروهی به‌جای QR، اولین عکس محصول را از صفحه لینک دانلود کن")
    return parser


def main() -> int:
    parser = build_arg_parser()
    args = parser.parse_args()

    if len(sys.argv) > 1:
        return run_cli(args)

    if tk is None:
        parser.print_help()
        return 1

    root = tk.Tk()
    app = QRApp(root)
    root.mainloop()
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
